"""XLSX / XLS extractor.

Each worksheet -> one Page rendered as a Markdown table. Cells are
treated as text — XLSX rarely contains scanned images, so OCR is only
triggered for embedded images if present (openpyxl exposes _images).

.xls files are read via pandas + xlrd.
"""
from __future__ import annotations

import io
import logging
from pathlib import Path
from typing import List

import openpyxl
import pandas as pd
from PIL import Image

from .base import BaseExtractor, Block, ExtractedDoc, Page

log = logging.getLogger(__name__)


class XlsxExtractor(BaseExtractor):
    extensions = ["xlsx", "xlsm", "xls"]
    format_name = "xlsx"

    def extract(self, path: str | Path) -> ExtractedDoc:
        path = Path(path)
        suf = path.suffix.lower()
        if suf == ".xls":
            return self._extract_xls(path)

        max_rows = int(self.cfg.extractors.xlsx.get("max_rows_per_sheet", 5000))
        max_cols = int(self.cfg.extractors.xlsx.get("max_cols_per_sheet", 200))
        min_side = int(self.cfg.ocr.min_image_side_px)

        wb = openpyxl.load_workbook(str(path), data_only=True, read_only=False)
        doc = ExtractedDoc(source_path=str(path), format="xlsx")

        for i, name in enumerate(wb.sheetnames):
            ws = wb[name]
            page = Page(index=i, label=f"Sheet: {name}")
            md = self._sheet_to_md(ws, max_rows, max_cols)
            if md:
                page.blocks.append(Block(type="table", text=md,
                                         source=f"sheet:{name}"))

            # Embedded images (charts, logos)
            for img in self._sheet_images(ws, min_side):
                page.blocks.append(Block(type="image", image=img,
                                         source=f"sheet:{name}:image",
                                         meta={"w": img.width, "h": img.height}))
            doc.pages.append(page)

        return doc

    # ------------------------------------------------------------------

    def _extract_xls(self, path: Path) -> ExtractedDoc:
        if not self.cfg.extractors.xlsx.get("legacy_xls_via_pandas", True):
            raise RuntimeError(".xls disabled via config")
        max_rows = int(self.cfg.extractors.xlsx.get("max_rows_per_sheet", 5000))
        max_cols = int(self.cfg.extractors.xlsx.get("max_cols_per_sheet", 200))
        sheets = pd.read_excel(str(path), sheet_name=None, engine="xlrd")
        doc = ExtractedDoc(source_path=str(path), format="xls")
        for i, (name, df) in enumerate(sheets.items()):
            df = df.iloc[:max_rows, :max_cols].fillna("")
            page = Page(index=i, label=f"Sheet: {name}")
            md = df.to_markdown(index=False) if not df.empty else ""
            if md:
                page.blocks.append(Block(type="table", text=md,
                                         source=f"sheet:{name}"))
            doc.pages.append(page)
        return doc

    # ------------------------------------------------------------------

    @staticmethod
    def _sheet_to_md(ws, max_rows: int, max_cols: int) -> str:
        rows: List[List[str]] = []
        for r, row in enumerate(ws.iter_rows(values_only=True)):
            if r >= max_rows:
                break
            row = row[:max_cols]
            rows.append(["" if v is None else str(v).replace("|", "\\|").replace("\n", " ")
                         for v in row])
        # Trim fully empty trailing rows
        while rows and not any(cell.strip() for cell in rows[-1]):
            rows.pop()
        if not rows:
            return ""
        width = max(len(r) for r in rows)
        rows = [r + [""] * (width - len(r)) for r in rows]
        header = rows[0] if any(c.strip() for c in rows[0]) else [f"col{i+1}" for i in range(width)]
        body = rows[1:] if rows[0] is header else rows
        md = ["| " + " | ".join(header) + " |",
              "| " + " | ".join(["---"] * width) + " |"]
        for r in body:
            md.append("| " + " | ".join(r) + " |")
        return "\n".join(md)

    @staticmethod
    def _sheet_images(ws, min_side: int) -> List[Image.Image]:
        out: List[Image.Image] = []
        for image in getattr(ws, "_images", []):
            try:
                data = image._data() if callable(getattr(image, "_data", None)) else None
                if not data:
                    continue
                img = Image.open(io.BytesIO(data)).convert("RGB")
            except Exception as e:
                log.debug("Skipping worksheet image: %s", e)
                continue
            if img.width >= min_side and img.height >= min_side:
                out.append(img)
        return out
